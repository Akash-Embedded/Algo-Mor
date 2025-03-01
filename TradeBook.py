import pandas as pd
from collections import namedtuple
import SMART_API_CONSTANT as CONST
import openpyxl
from openpyxl import load_workbook

class trade_book:
    def __init__(self, file_path: str, sheet_name):
        self.file_path = file_path
        self.stocks = []
        self.file_path = file_path
        self.read_stock_names()
        self.sheet_name = sheet_name
        # Load existing sheet
        self.existing_data = pd.read_excel(file_path, sheet_name=sheet_name)

    def read_stock_names(self):
        # Read the Excel file into a DataFrame
        try:
            df = pd.read_excel(self.file_path, usecols="A")  # Only read column A
            df.dropna(inplace=True)  # Remove empty rows

            # Convert the column 'A' into a list
            self.stocks = df.iloc[:, 0].tolist()
        except Exception as e:
            print(f"Error reading the Excel file: {e}")

    def append_to_excel(self, file_path, df, sheet_name):
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)

            # Check if the sheet exists
            if sheet_name in workbook.sheetnames:
                with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    # Get the last row of the existing sheet
                    existing_sheet = workbook[sheet_name]
                    startrow = existing_sheet.max_row

                    # Append data starting from the next available row
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow, header=False)
            else:
                # If the sheet doesn't exist, create it
                with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        except FileNotFoundError:
            # If the file doesn't exist, create it and write the data
            with pd.ExcelWriter(file_path, mode='w', engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def write_high_volume_stock(self, candle, excel, sheet_name, stock_name, volume_20_ma, price_diff_percentage):
        candle

        # Combine candle and technical data into a DataFrame
        data = []
        data.append({
                "name": stock_name,
                "volume_change": ((candle.volume - volume_20_ma) / volume_20_ma ) * 100,
                "price_change": price_diff_percentage,
                "date":  candle.time.replace(tzinfo=None),  # Remove timezone info
                "open": candle.open,
                "high": candle.high,
                "low": candle.low,
                "close": candle.close,
                "volume": candle.volume,
                "volume_20_ma": volume_20_ma,
                "rsi": candle.rsi,  # Placeholder if RSI is not calculated
            })

        df = pd.DataFrame(data)
        self.append_to_excel(excel, df, sheet_name)

    def update_indicators_sheet(self, candles, excel_path, stock, mor_dict, short_dict):
        current_candle = candles[-1]
        past_candle = candles[-2]
        data = [{
            "name": stock,
            "volume_change": ((current_candle.volume - current_candle.volume_20_ma) / current_candle.volume_20_ma) * 100,
            "price_change": ((current_candle.close - past_candle.close) / past_candle.close) * 100,
            "date": current_candle.time.replace(tzinfo=None),  # Remove timezone info
            "open": current_candle.open,
            "high": current_candle.high,
            "low": current_candle.low,
            "close": current_candle.close,
            "volume": current_candle.volume,
            "volume_20_ma": current_candle.volume_20_ma,
            "rsi": current_candle.rsi,
            "mor": mor_dict["indicator"],
            "mor_age": mor_dict["age"],
            "mor_date": mor_dict["time"].replace(tzinfo=None),
            "price_above_200_sma": mor_dict["ma_200_cross"],
        }]
        new_data = pd.DataFrame(data)
    
        try:
        #if True:
    
            # Ensure matching columns between existing_data and new_data
            for col in ["volume_change", "price_change", "date", "open", "high", "low", "close", "volume", "volume_20_ma", "rsi", "mor", "mor_age", "mor_date", "price_above_200_sma"]:
                if col in self.existing_data.columns and col in new_data.columns:
                    new_data[col] = new_data[col].astype(self.existing_data[col].dtype, errors="ignore")
    
            # Update matching stock row or append if it doesn't exist
            if stock in self.existing_data["name"].values:
                # Replace matching columns with new data
                for col in ["volume_change", "price_change", "date", "open", "high", "low", "close", "volume", "volume_20_ma", "rsi", "mor", "mor_age", "mor_date", "price_above_200_sma"]:
                    self.existing_data.loc[self.existing_data["name"] == stock, col] = new_data[col].values[0]
            else:
                # Append new row for non-existing stock
                self.existing_data = pd.concat([self.existing_data, new_data], ignore_index=True)
    
        except Exception as e:
            print(f"Error updating indicators sheet: {e}")

    def calculate_monthly_ohlc(self, candles, stock):
        df = pd.DataFrame(candles)
    
        # Ensure datetime format and remove timezone
        df["time"] = pd.to_datetime(df["time"]).dt.tz_localize(None)
        df.sort_values("time", inplace=True)
    
        # Identify last Thursday of each month
        df["year_month"] = df["time"].dt.to_period("M")
        
        # Find the last day of the month for each row
        df["last_day_of_month"] = df["time"].dt.to_period("M").dt.to_timestamp() + pd.offsets.MonthEnd(0)
    
        # Find the last Thursday of the month
        df["last_thursday"] = df["last_day_of_month"] - pd.to_timedelta((df["last_day_of_month"].dt.weekday - 3) % 7, unit='D')
        
        # Identify if the current row is the last Thursday of the month
        df["is_last_thursday"] = df["time"] == df["last_thursday"]
    
        # Mark month changes
        df["month_group"] = df["is_last_thursday"].shift(1).fillna(False).cumsum()
    
        # Aggregate to get monthly OHLC and percentage change
        monthly_df = df.groupby("month_group").agg(
            open=("open", "first"),
            high=("high", "max"),
            low=("low", "min"),
            close=("close", "last"),
            start_date=("time", "first"),
            end_date=("time", "last"),
        )
    
        # Remove timezone from start_date and end_date before writing to Excel
        monthly_df["start_date"] = monthly_df["start_date"].dt.tz_localize(None)
        monthly_df["end_date"] = monthly_df["end_date"].dt.tz_localize(None)
    
        # Calculate percentage move
        monthly_df["monthly_percentage_move"] = (
            (monthly_df["close"] - monthly_df["open"]) / monthly_df["open"] * 100
        )
    
        # Add stock name to each entry
        monthly_df["stock"] = stock  
    
        new_data = monthly_df.reset_index(drop=True)
        self.existing_data = pd.concat([self.existing_data, new_data], ignore_index=True)

    def calculate_weekly_ohlc(self, candles, stock):
        df = pd.DataFrame(candles)
    
        # Ensure datetime format and remove timezone
        df["time"] = pd.to_datetime(df["time"]).dt.tz_localize(None)
        df.sort_values("time", inplace=True)
    
        # Define a custom week grouping (Week starts on Monday and ends on Thursday)
        df["week_start"] = df["time"] - pd.to_timedelta((df["time"].dt.weekday - 0) % 7, unit="D")
    
        # Add an extra column to explicitly mark Thursday as the end of the week
        df["week_end"] = df["week_start"] + pd.Timedelta(days=3)  # Week ends 3 days after Monday
    
        # Aggregate OHLC values based on week start
        weekly_df = df.groupby("week_start").agg(
            open=("open", "first"),
            high=("high", "max"),
            low=("low", "min"),
            close=("close", "last"),
            start_date=("time", "first"),
            end_date=("time", "last"),
        )
    
        # Add week_end explicitly as Thursday
        weekly_df["end_date"] = weekly_df["start_date"] + pd.Timedelta(days=3)  # End date explicitly set as Thursday
    
        # Remove timezone from start_date and end_date
        weekly_df["start_date"] = weekly_df["start_date"].dt.tz_localize(None)
        weekly_df["end_date"] = weekly_df["end_date"].dt.tz_localize(None)
    
        # Calculate weekly percentage move
        weekly_df["weekly_percentage_move"] = (
            (weekly_df["close"] - weekly_df["open"]) / weekly_df["open"] * 100
        )
    
        # Add stock name to each entry
        weekly_df["stock"] = stock  
    
        # Reset index and add new data to the existing data
        new_data = weekly_df.reset_index(drop=True)
        self.existing_data = pd.concat([self.existing_data, new_data], ignore_index=True)


    def read_high_volume_stock(self, excel, sheet_name):
        try:
            # Read the data from the specified sheet in the Excel file
            df = pd.read_excel(excel, sheet_name=sheet_name)
            return df.to_dict(orient='records')

        except FileNotFoundError:
            print(f"The file {excel} was not found.")
            return None
        except ValueError:
            print(f"The sheet {sheet_name} does not exist in the file {excel}.")
            return None

    def get_stocks(self):
        return self.stocks
    
    def write_sheet(self):
        # Write the updated data back to Excel
        with pd.ExcelWriter(self.file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            self.existing_data.to_excel(writer, self.sheet_name, index=False)


    def update_backtrace_sheet(self, stock, phase, entry_candle, exit_candle, mor_age, profit_loss):
        data = [{
            "name": stock,
            "phase": phase, 
            "age" : mor_age,
            "entry_price":getattr(entry_candle, 'close'),
            "entry_date": getattr(entry_candle, 'time').replace(tzinfo=None),
            "exit_price": getattr(exit_candle, 'close'),
            "exit_date":  getattr(exit_candle, 'time').replace(tzinfo=None),
            "PL": profit_loss,
            "entry_rsi":getattr(entry_candle, 'rsi'),
            "exit_rsi": getattr(exit_candle, 'rsi'),
        }]
        new_data = pd.DataFrame(data)
        self.existing_data = pd.concat([self.existing_data, new_data], ignore_index=True)

    def update_backtrace_summry(self, total_profit_loss):
        data = [{
            "total_profit_loss": total_profit_loss,
        }]
        new_data = pd.DataFrame(data)
        self.existing_data = pd.concat([self.existing_data, new_data], ignore_index=True)

    def summerised_monthly_profit(self, date_column_name, profit_column_name, stock_column_name):
        # Ensure the DateTime column is in datetime format
        self.existing_data['DateTime'] = pd.to_datetime(self.existing_data[date_column_name])

        # Extract year and month
        self.existing_data['Year-Month'] = self.existing_data['DateTime'].dt.to_period('M')

        # Group by Year-Month and calculate total profit/loss
        summary = self.existing_data.groupby('Year-Month')[profit_column_name].sum().reset_index()

        # Convert Year-Month back to string for better Excel compatibility
        summary['Year-Month'] = summary['Year-Month'].astype(str)

        # Save the summary to a new sheet in the same workbook
        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            summary.to_excel(writer, sheet_name="summary-total", index=False)

        # Group by Year-Month and calculate total profit/loss
        summary = self.existing_data.groupby(['Year-Month', stock_column_name])[profit_column_name].sum().reset_index()

        # Convert Year-Month back to string for better Excel compatibility
        summary['Year-Month'] = summary['Year-Month'].astype(str)

        # Save the summary to a new sheet in the same workbook
        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            summary.to_excel(writer, sheet_name="summary-stock", index=False)

        # Group by Year-Month and calculate total profit/loss
        summary = self.existing_data.groupby(stock_column_name)[profit_column_name].sum().reset_index()

        # Convert Year-Month back to string for better Excel compatibility
        summary[stock_column_name] = summary[stock_column_name].astype(str)

        # Save the summary to a new sheet in the same workbook
        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            summary.to_excel(writer, sheet_name="summary-stockWise", index=False)


        # Group by stock and count number of profit and loss trades
        profit_loss_summary = self.existing_data.copy()
        profit_loss_summary['TradeType'] = profit_loss_summary[profit_column_name].apply(lambda x: 'Profit' if x > 0 else 'Loss')

        # Count profit and loss trades per stock
        profit_loss_count = profit_loss_summary.groupby([stock_column_name, 'TradeType']).size().unstack(fill_value=0).reset_index()

        # Save the profit/loss count to a new sheet
        with pd.ExcelWriter(self.file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            profit_loss_count.to_excel(writer, sheet_name="summary-tradeCounts", index=False)

    def empty_Sheet(self):
        self.existing_data = pd.DataFrame()


